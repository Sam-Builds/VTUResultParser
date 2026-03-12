
import os
import queue
import sys
import threading
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox
import tkinter as tk
from tkinter import ttk

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from tkinterdnd2 import DND_FILES, TkinterDnD

sys.path.insert(0, str(Path(__file__).parent))
from pdfparser import parse_scanned_vtu

W, H         = 700, 620
BG           = "#f5f7fa"
CARD         = "#ffffff"
ACCENT       = "#1a73e8"
ACCENT_DK    = "#1558b0"
ACCENT_PALE  = "#e8f0fe"
ACCENT_HOVER = "#d2e3fc"
T_DARK       = "#202124"
T_MID        = "#5f6368"
T_LIGHT      = "#9aa0a6"
BORDER_CLR   = "#dadce0"
OK_CLR       = "#1e8e3e"
ERR_CLR      = "#d93025"
WARN_CLR     = "#f9ab00"

F_TITLE  = ("Segoe UI", 17, "bold")
F_HEAD   = ("Segoe UI", 11, "bold")
F_BODY   = ("Segoe UI", 10)
F_SMALL  = ("Segoe UI", 9)
F_LOG    = ("Consolas", 9)

COLS = ["USN", "Name", "Subject Code", "Subject Name",
        "Internal", "External", "Total", "Result"]

INST_NAME      = "YENEPOYA INSTITUTE OF TECHNOLOGY"
DEPT_NAME      = "Department of CSE(Data Science)"
SHEET_TITLE    = "RESULT SHEET 2025-26 – Before Revaluation – I Sem"
CREDIT_SHEET_TITLE = "RESULT SHEET 2025-26 – After Revaluation – Credit Points – I Sem"
MAX_SUBJ_MARKS = 100  
PASS_MARK      = 40   



def _thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _save_excel(rows: list, out_dir: Path, subject_credits: dict[str, int]) -> Path:
    sorted_rows = sorted(rows, key=lambda r: (r.get("USN", ""), r.get("Subject Code", "")))
    student_order = []
    student_names = {}
    student_subjs = {}      
    for r in sorted_rows:
        usn  = r.get("USN", "")
        code = r.get("Subject Code", "")
        if usn not in student_subjs:
            student_order.append(usn)
            student_names[usn] = r.get("Name", "")
            student_subjs[usn] = {}
        student_subjs[usn][code] = r

    all_codes = sorted({r.get("Subject Code", "") for r in rows if r.get("Subject Code", "")})
    n_subj    = len(all_codes)
    cl = get_column_letter
    COL_SNO = 1
    COL_USN = 2
    COL_NAME = 3
    white_fill  = PatternFill("solid", fgColor="FFFFFF")

    hdr_font    = Font(bold=True, color="000000", name="Segoe UI", size=9)
    sub_font    = Font(bold=True, color="000000", name="Segoe UI", size=8)
    data_font   = Font(name="Segoe UI", size=9)
    title_font  = Font(name="Segoe UI", size=11, bold=True, color="000000")

    hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ctr_align  = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin       = _thin_border()

    wb = Workbook()

    def _write_title_block(ws, title_text: str, n_cols: int):
        for r_i, (text, fsize, row_h) in enumerate([
            (INST_NAME, 14, 24),
            (DEPT_NAME, 11, 20),
            (title_text, 11, 20),
        ], 1):
            ws.merge_cells(start_row=r_i, start_column=1, end_row=r_i, end_column=n_cols)
            c = ws.cell(r_i, 1, text)
            c.fill = white_fill
            c.font = Font(name="Segoe UI", size=fsize, bold=True, color="000000")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin
            ws.row_dimensions[r_i].height = row_h

    def _write_summary_block(ws, data_row: int, last_dr: int, tot_cols: list[int], col_fail: int):
        summary_start = last_dr + 3
        summary_labels = [
            "SUBJECT",
            "NO. OF STUDENTS TAKING THE EXAM",
            "NO. OF STUDENTS PASS IN THE PAPER",
            "NO. OF STUDENTS FAIL IN THE PAPER",
            "RESULT IN PERCENTAGE (%)",
            "NO. OF STUDENTS PASSED IN ALL SUBJECTS",
            "NO OF STUDENTS FAIL",
            "RESULT IN PERCENTAGE (%)",
        ]

        summary_col0 = 4
        for idx, label in enumerate(summary_labels):
            row_no = summary_start + idx
            ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=3)
            c = ws.cell(row_no, 1, label)
            c.fill = white_fill
            c.font = hdr_font if idx == 0 else title_font
            c.alignment = left_align if idx > 0 else hdr_align
            c.border = thin
            for extra_col in range(2, 4):
                ws.cell(row_no, extra_col).fill = white_fill
                ws.cell(row_no, extra_col).border = thin

        for i, code in enumerate(all_codes):
            col = summary_col0 + i
            c = ws.cell(summary_start, col, code)
            c.fill = white_fill
            c.font = hdr_font
            c.alignment = ctr_align
            c.border = thin

            tot_col = tot_cols[i]
            tot_rng = f"{cl(tot_col)}{data_row}:{cl(tot_col)}{last_dr}"
            ws.cell(summary_start + 1, col, f"=COUNT({tot_rng})")
            ws.cell(summary_start + 2, col, f'=COUNTIF({tot_rng},">={PASS_MARK}")')
            ws.cell(summary_start + 3, col, f'=COUNTIF({tot_rng},"<{PASS_MARK}")')
            ws.cell(summary_start + 4, col,
                    f"=IF({cl(col)}{summary_start + 1}=0,0,{cl(col)}{summary_start + 2}/{cl(col)}{summary_start + 1}*100)")

            for row_no in range(summary_start + 1, summary_start + 5):
                cell = ws.cell(row_no, col)
                cell.fill = white_fill
                cell.font = data_font
                cell.alignment = ctr_align
                cell.border = thin
            ws.cell(summary_start + 4, col).number_format = "0.00"

        total_students = len(student_order)
        overall_subject_span_end = summary_col0 + max(n_subj - 1, 0)
        if n_subj:
            ws.merge_cells(start_row=summary_start + 5, start_column=summary_col0,
                           end_row=summary_start + 5, end_column=overall_subject_span_end)
            ws.merge_cells(start_row=summary_start + 6, start_column=summary_col0,
                           end_row=summary_start + 6, end_column=overall_subject_span_end)
            ws.merge_cells(start_row=summary_start + 7, start_column=summary_col0,
                           end_row=summary_start + 7, end_column=overall_subject_span_end)

        ws.cell(summary_start + 5, summary_col0,
                f'=COUNTIF({cl(col_fail)}{data_row}:{cl(col_fail)}{last_dr},0)')
        ws.cell(summary_start + 6, summary_col0,
                f'=COUNTIF({cl(col_fail)}{data_row}:{cl(col_fail)}{last_dr},">0")')
        ws.cell(summary_start + 7, summary_col0,
                f"=IF({total_students}=0,0,{cl(summary_col0)}{summary_start + 5}/{total_students}*100)")

        for row_no in range(summary_start + 5, summary_start + 8):
            c = ws.cell(row_no, summary_col0)
            c.fill = white_fill
            c.font = data_font
            c.alignment = ctr_align
            c.border = thin
            if row_no == summary_start + 7:
                c.number_format = "0.00"
            for col in range(summary_col0 + 1, overall_subject_span_end + 1):
                ws.cell(row_no, col).fill = white_fill
                ws.cell(row_no, col).border = thin

    ws = wb.active
    ws.title = "Result Sheet"
    DATA_ROW = 6
    COL_S0 = 4
    col_total = COL_S0 + n_subj * 3
    col_pct = col_total + 1
    col_fail = col_pct + 1
    n_cols = col_fail

    _write_title_block(ws, SHEET_TITLE, n_cols)

    for col, text in [(COL_SNO, "S.No"), (COL_USN, "USN"), (COL_NAME, "STUDENT NAME")]:
        ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
        c = ws.cell(4, col, text)
        c.fill = white_fill; c.font = hdr_font; c.alignment = hdr_align; c.border = thin

    for i, code in enumerate(all_codes):
        c0 = COL_S0 + i * 3
        ws.merge_cells(start_row=4, start_column=c0, end_row=4, end_column=c0 + 2)
        c = ws.cell(4, c0, code)
        c.fill = white_fill; c.font = hdr_font; c.alignment = hdr_align; c.border = thin

    for col, text in [
        (col_total, "Total\nMarks"),
        (col_pct, "%"),
        (col_fail, "Total no.\nof Fail"),
    ]:
        ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
        c = ws.cell(4, col, text)
        c.fill = white_fill; c.font = hdr_font; c.alignment = hdr_align; c.border = thin
    ws.row_dimensions[4].height = 30

    for i in range(n_subj):
        c0 = COL_S0 + i * 3
        for off, lbl in enumerate(("INT", "EXT", "TOT")):
            c = ws.cell(5, c0 + off, lbl)
            c.fill = white_fill; c.font = sub_font; c.alignment = ctr_align; c.border = thin
    ws.row_dimensions[5].height = 16

    tot_cols = [COL_S0 + i * 3 + 2 for i in range(n_subj)]

    for s_idx, usn in enumerate(student_order):
        dr = DATA_ROW + s_idx

        def wc(col, val, align=ctr_align, fmt=None, _dr=dr):
            c = ws.cell(_dr, col, val)
            c.fill = white_fill
            c.font = data_font
            c.alignment = align
            c.border = thin
            if fmt:
                c.number_format = fmt
            return c

        wc(COL_SNO, s_idx + 1)
        wc(COL_USN, usn)
        wc(COL_NAME, student_names.get(usn, ""), left_align)

        for i, code in enumerate(all_codes):
            subj = student_subjs[usn].get(code, {})
            c0 = COL_S0 + i * 3
            for off, key in enumerate(("Internal", "External", "Total")):
                raw = subj.get(key, "")
                try:
                    val = int(raw) if raw != "" else ""
                except (ValueError, TypeError):
                    val = raw
                wc(c0 + off, val)

        wc(col_total, "=" + "+".join(f"{cl(c)}{dr}" for c in tot_cols), fmt="0")
        wc(col_pct, f"={cl(col_total)}{dr}/{n_subj * MAX_SUBJ_MARKS}*100", fmt="0.00")
        fail_expr = "+".join(
            f"ISNUMBER({cl(c)}{dr})*({cl(c)}{dr}<{PASS_MARK})" for c in tot_cols
        )
        wc(col_fail, f"={fail_expr}")
        ws.row_dimensions[dr].height = 16

    last_dr = DATA_ROW + max(len(student_order) - 1, 0)
    _write_summary_block(ws, DATA_ROW, last_dr, tot_cols, col_fail)

    ws.column_dimensions[cl(COL_SNO)].width = 5
    ws.column_dimensions[cl(COL_USN)].width = 14
    ws.column_dimensions[cl(COL_NAME)].width = 30
    for i in range(n_subj):
        c0 = COL_S0 + i * 3
        ws.column_dimensions[cl(c0)].width = 5
        ws.column_dimensions[cl(c0 + 1)].width = 5
        ws.column_dimensions[cl(c0 + 2)].width = 5
    ws.column_dimensions[cl(col_total)].width = 10
    ws.column_dimensions[cl(col_pct)].width = 7
    ws.column_dimensions[cl(col_fail)].width = 10
    ws.freeze_panes = f"D{DATA_ROW}"

    ws_credit = wb.create_sheet("Credit Sheet")
    C_DATA_ROW = 6
    C_COL_S0 = 4
    c_col_total_cp = C_COL_S0 + n_subj * 3
    c_col_sgpa = c_col_total_cp + 1
    c_col_remarks = c_col_sgpa + 1
    c_col_backlogs = c_col_remarks + 1
    c_col_pct = c_col_backlogs + 1
    c_n_cols = c_col_pct
    total_credits = sum(subject_credits.get(code, 0) for code in all_codes)

    _write_title_block(ws_credit, CREDIT_SHEET_TITLE, c_n_cols)

    for col, text in [(COL_SNO, "S.No"), (COL_USN, "USN"), (COL_NAME, "STUDENT NAME")]:
        ws_credit.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
        c = ws_credit.cell(4, col, text)
        c.fill = white_fill; c.font = hdr_font; c.alignment = hdr_align; c.border = thin

    for i, code in enumerate(all_codes):
        c0 = C_COL_S0 + i * 3
        ws_credit.merge_cells(start_row=4, start_column=c0, end_row=4, end_column=c0 + 2)
        c = ws_credit.cell(4, c0, code)
        c.fill = white_fill; c.font = hdr_font; c.alignment = hdr_align; c.border = thin

    for col, text in [
        (c_col_total_cp, "TOTAL CP"),
        (c_col_sgpa, "SGPA"),
        (c_col_remarks, "REMARKS"),
        (c_col_backlogs, "No. of Backlogs"),
        (c_col_pct, "Percentage"),
    ]:
        ws_credit.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
        c = ws_credit.cell(4, col, text)
        c.fill = white_fill; c.font = hdr_font; c.alignment = hdr_align; c.border = thin
    ws_credit.row_dimensions[4].height = 30

    for i in range(n_subj):
        c0 = C_COL_S0 + i * 3
        for off, lbl in enumerate(("TOT", "GP", "CP")):
            c = ws_credit.cell(5, c0 + off, lbl)
            c.fill = white_fill; c.font = sub_font; c.alignment = ctr_align; c.border = thin
    ws_credit.row_dimensions[5].height = 16

    credit_tot_cols = [C_COL_S0 + i * 3 for i in range(n_subj)]
    credit_gp_cols = [C_COL_S0 + i * 3 + 1 for i in range(n_subj)]
    credit_cp_cols = [C_COL_S0 + i * 3 + 2 for i in range(n_subj)]

    for s_idx, usn in enumerate(student_order):
        dr = C_DATA_ROW + s_idx

        def wc_credit(col, val, align=ctr_align, fmt=None, _dr=dr):
            c = ws_credit.cell(_dr, col, val)
            c.fill = white_fill
            c.font = data_font
            c.alignment = align
            c.border = thin
            if fmt:
                c.number_format = fmt
            return c

        wc_credit(COL_SNO, s_idx + 1)
        wc_credit(COL_USN, usn)
        wc_credit(COL_NAME, student_names.get(usn, ""), left_align)

        for i, code in enumerate(all_codes):
            subj = student_subjs[usn].get(code, {})
            c0 = C_COL_S0 + i * 3
            raw_total = subj.get("Total", "")
            try:
                total_val = int(raw_total) if raw_total != "" else ""
            except (ValueError, TypeError):
                total_val = raw_total

            wc_credit(c0, total_val)

            tot_ref = f"{cl(c0)}{dr}"
            gp_ref = f"{cl(c0 + 1)}{dr}"
            credit = int(subject_credits.get(code, 0))
            gp_formula = (
                f'=IF(ISNUMBER({tot_ref}),IF({tot_ref}>=90,10,IF({tot_ref}>=80,9,'
                f'IF({tot_ref}>=70,8,IF({tot_ref}>=60,7,IF({tot_ref}>=55,6,'
                f'IF({tot_ref}>=50,5,IF({tot_ref}>=40,4,0))))))),"")'
            )
            wc_credit(c0 + 1, gp_formula)
            wc_credit(c0 + 2, f'=IF({gp_ref}="","",{gp_ref}*{credit})')

        wc_credit(c_col_total_cp,
                  "=" + "+".join(f"{cl(c)}{dr}" for c in credit_cp_cols),
                  fmt="0")
        wc_credit(c_col_sgpa,
                  f"=IF({total_credits}=0,0,{cl(c_col_total_cp)}{dr}/{total_credits})",
                  fmt="0.00")

        backlog_expr = "+".join(
            f"ISNUMBER({cl(c)}{dr})*({cl(c)}{dr}<{PASS_MARK})" for c in credit_tot_cols
        )
        wc_credit(c_col_backlogs, f"={backlog_expr}")
        wc_credit(c_col_remarks,
                  f'=IF({cl(c_col_backlogs)}{dr}=0,"PASS","FAIL")')

        tot_sum_expr = "+".join(f"{cl(c)}{dr}" for c in credit_tot_cols)
        wc_credit(c_col_pct,
                  f"=({tot_sum_expr})/{n_subj * MAX_SUBJ_MARKS}*100",
                  fmt="0.00")
        ws_credit.row_dimensions[dr].height = 16

    c_last_dr = C_DATA_ROW + max(len(student_order) - 1, 0)
    _write_summary_block(ws_credit, C_DATA_ROW, c_last_dr, credit_tot_cols, c_col_backlogs)

    ws_credit.column_dimensions[cl(COL_SNO)].width = 5
    ws_credit.column_dimensions[cl(COL_USN)].width = 14
    ws_credit.column_dimensions[cl(COL_NAME)].width = 30
    for i in range(n_subj):
        c0 = C_COL_S0 + i * 3
        ws_credit.column_dimensions[cl(c0)].width = 5
        ws_credit.column_dimensions[cl(c0 + 1)].width = 5
        ws_credit.column_dimensions[cl(c0 + 2)].width = 5
    ws_credit.column_dimensions[cl(c_col_total_cp)].width = 10
    ws_credit.column_dimensions[cl(c_col_sgpa)].width = 7
    ws_credit.column_dimensions[cl(c_col_remarks)].width = 10
    ws_credit.column_dimensions[cl(c_col_backlogs)].width = 14
    ws_credit.column_dimensions[cl(c_col_pct)].width = 10
    ws_credit.freeze_panes = f"D{C_DATA_ROW}"

    ws2 = wb.create_sheet("Raw Data")
    RAW_COLS = ["USN", "Name", "Subject Code", "Subject Name",
                "Internal", "External", "Total", "Result"]
    raw_widths = {
        "USN": 16, "Name": 26, "Subject Code": 14, "Subject Name": 42,
        "Internal": 10, "External": 10, "Total": 8, "Result": 8,
    }
    for ci, cn in enumerate(RAW_COLS, 1):
        c = ws2.cell(1, ci, cn)
        c.fill = white_fill
        c.font = Font(bold=True, color="000000", name="Segoe UI", size=10)
        c.alignment = hdr_align; c.border = thin
    ws2.row_dimensions[1].height = 28

    data_font2 = Font(name="Segoe UI", size=10)
    for ri, r in enumerate(sorted_rows, 1):
        rf = white_fill
        for ci, cn in enumerate(RAW_COLS, 1):
            c = ws2.cell(ri + 1, ci, r.get(cn, ""))
            c.fill = rf; c.font = data_font2; c.border = thin
            c.alignment = left_align if cn in ("Name", "Subject Name") else ctr_align
        ws2.row_dimensions[ri + 1].height = 16

    for ci, cn in enumerate(RAW_COLS, 1):
        ws2.column_dimensions[cl(ci)].width = raw_widths[cn]
    ws2.freeze_panes = "A2"

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out   = out_dir / f"VTU_Results_{stamp}.xlsx"
    try:
        wb.save(out)
    except PermissionError:
        out = out_dir / f"VTU_Results_{datetime.now().strftime('%Y%m%d_%H%M%S')}_new.xlsx"
        wb.save(out)

    return out



class VTUParserApp(TkinterDnD.Tk):

    def __init__(self):
        super().__init__()
        self.title("VTU Marksheet Bulk Parser")
        self.geometry(f"{W}x{H}")
        self.resizable(False, False)
        self.configure(bg=BG)
        self.update_idletasks()
        x = (self.winfo_screenwidth()  - W) // 2
        y = (self.winfo_screenheight() - H) // 2
        self.geometry(f"{W}x{H}+{x}+{y}")

        self._folder: Path | None = None
        self._busy = False
        self._q: queue.Queue = queue.Queue()
        self._last_credits: dict[str, int] = {}

        self._build_ui()
        self._poll_queue()


    def _build_ui(self):
        hdr = tk.Frame(self, bg=ACCENT, height=60)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="VTU Marksheet Bulk Parser",
                 font=F_TITLE, bg=ACCENT, fg="white").pack(side="left", padx=20)
    

        body = tk.Frame(self, bg=BG)
        body.pack(fill="both", expand=True, padx=24, pady=16)
        self._zone = tk.Frame(body, bg=ACCENT_PALE,
                              highlightbackground=ACCENT, highlightthickness=2)
        self._zone.pack(fill="x", ipady=16)

        tk.Label(self._zone, text="📂", font=("Segoe UI Emoji", 30),
                 bg=ACCENT_PALE, fg=ACCENT).pack(pady=(12, 2))
        tk.Label(self._zone, text="Drop a folder here",
                 font=F_HEAD, bg=ACCENT_PALE, fg=ACCENT).pack()
        tk.Label(self._zone, text="containing your VTU PDF marksheets",
                 font=F_SMALL, bg=ACCENT_PALE, fg=T_MID).pack(pady=(0, 12))

        self._zone.drop_target_register(DND_FILES)
        self._zone.dnd_bind("<<DragEnter>>", self._drag_enter)
        self._zone.dnd_bind("<<DragLeave>>", self._drag_leave)
        self._zone.dnd_bind("<<Drop>>",      self._on_drop)
        for child in self._zone.winfo_children():
            child.drop_target_register(DND_FILES)
            child.dnd_bind("<<DragEnter>>", self._drag_enter)
            child.dnd_bind("<<DragLeave>>", self._drag_leave)
            child.dnd_bind("<<Drop>>",      self._on_drop)

        ctrl = tk.Frame(body, bg=BG)
        ctrl.pack(fill="x", pady=(12, 0))

        self._btn_browse = tk.Button(
            ctrl, text="  Browse Folder  ", font=F_HEAD,
            bg=ACCENT, fg="white", activebackground=ACCENT_DK,
            activeforeground="white", bd=0, padx=16, pady=8,
            relief="flat", cursor="hand2", command=self._browse
        )
        self._btn_browse.pack(side="left")

        info = tk.Frame(ctrl, bg=BG)
        info.pack(side="left", padx=14)
        self._lbl_path  = tk.Label(info, text="No folder selected.",
                                   font=F_BODY, bg=BG, fg=T_MID,
                                   anchor="w", wraplength=430, justify="left")
        self._lbl_path.pack(anchor="w")
        self._lbl_count = tk.Label(info, text="",
                                   font=F_SMALL, bg=BG, fg=T_MID, anchor="w")
        self._lbl_count.pack(anchor="w")

        ttk.Separator(body, orient="horizontal").pack(fill="x", pady=12)

        self._progress_var = tk.DoubleVar(value=0)
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("Blue.Horizontal.TProgressbar",
                        troughcolor=BORDER_CLR, bordercolor=BORDER_CLR,
                        background=ACCENT, lightcolor=ACCENT, darkcolor=ACCENT_DK)
        pb_frame = tk.Frame(body, bg=BG)
        pb_frame.pack(fill="x")
        ttk.Progressbar(pb_frame, length=W - 52, mode="determinate",
                        maximum=100, variable=self._progress_var,
                        style="Blue.Horizontal.TProgressbar").pack(fill="x")

        log_outer = tk.Frame(body, bg=CARD,
                             highlightbackground=BORDER_CLR, highlightthickness=1)
        log_outer.pack(fill="both", expand=True, pady=(8, 0))

        self._log = tk.Text(log_outer, height=10, font=F_LOG, bg=CARD,
                            fg=T_DARK, bd=0, wrap="word",
                            state="disabled", cursor="arrow",
                            selectbackground=ACCENT_PALE)
        sb = ttk.Scrollbar(log_outer, command=self._log.yview)
        self._log.configure(yscrollcommand=sb.set)
        self._log.pack(side="left", fill="both", expand=True, padx=8, pady=6)
        sb.pack(side="right", fill="y")

        self._log.tag_configure("ok",   foreground=OK_CLR)
        self._log.tag_configure("err",  foreground=ERR_CLR)
        self._log.tag_configure("warn", foreground=WARN_CLR)
        self._log.tag_configure("info", foreground=T_MID)
        self._log.tag_configure("done", foreground=ACCENT,
                                font=("Segoe UI", 10, "bold"))

        btn_row = tk.Frame(body, bg=BG)
        btn_row.pack(pady=(10, 0))

        self._btn_parse = tk.Button(
            btn_row, text="  Parse & Export to Excel  ",
            font=("Segoe UI", 12, "bold"),
            bg=ACCENT, fg="white", activebackground=ACCENT_DK,
            activeforeground="white", bd=0, padx=28, pady=12,
            relief="flat", cursor="hand2", state="disabled",
            command=self._start_parse
        )
        self._btn_parse.pack()


    def _set_zone_bg(self, color):
        self._zone.configure(bg=color)
        for w in self._zone.winfo_children():
            w.configure(bg=color)

    def _drag_enter(self, event):
        self._set_zone_bg(ACCENT_HOVER)

    def _drag_leave(self, event):
        self._set_zone_bg(ACCENT_PALE)

    def _on_drop(self, event):
        self._set_zone_bg(ACCENT_PALE)
        try:
            paths = self.tk.splitlist(event.data.strip())
        except Exception:
            paths = [event.data.strip().strip("{}")]

        for raw in paths:
            path = Path(raw)
            if path.is_dir():
                self._set_folder(path)
                return
        self._log_msg("Dropped item is not a folder — please drop a folder.", "err")


    def _browse(self):
        folder = filedialog.askdirectory(
            title="Select folder containing VTU PDF marksheets"
        )
        if folder:
            self._set_folder(Path(folder))

    def _set_folder(self, path: Path):
        self._folder = path
        pdfs = sorted(path.glob("*.pdf"))
        self._lbl_path.configure(text=str(path), fg=T_DARK)
        if pdfs:
            self._lbl_count.configure(
                text=f"{len(pdfs)} PDF file(s) ready to parse.", fg=OK_CLR
            )
            self._btn_parse.configure(state="normal")
            self._log_msg(f"Folder: {path}", "info")
            names = ", ".join(p.name for p in pdfs[:6])
            suffix = f" … (+{len(pdfs) - 6} more)" if len(pdfs) > 6 else ""
            self._log_msg(f"PDFs found: {names}{suffix}", "info")
        else:
            self._lbl_count.configure(text="No PDF files found in this folder.", fg=ERR_CLR)
            self._btn_parse.configure(state="disabled")
            self._log_msg("No PDF files found in the selected folder.", "err")


    def _start_parse(self):
        if self._busy or not self._folder:
            return
        pdfs = sorted(self._folder.glob("*.pdf"))
        if not pdfs:
            messagebox.showwarning("No PDFs", "No PDF files found in the selected folder.")
            return
        self._busy = True
        self._btn_parse.configure(state="disabled", text="  Processing…  ")
        self._btn_browse.configure(state="disabled")
        self._progress_var.set(0)
        self._log_msg(f"\nStarting bulk parse of {len(pdfs)} file(s) …", "info")
        threading.Thread(target=self._worker,
                         args=(pdfs,), daemon=True).start()

    def _worker(self, pdfs: list):
        all_rows = []
        errors   = []
        n        = len(pdfs)

        for idx, pdf in enumerate(pdfs, 1):
            self._q.put(("log",      f"[{idx}/{n}]  {pdf.name}", "info"))
            self._q.put(("progress", int((idx - 1) / n * 100)))
            try:
                rows = parse_scanned_vtu(str(pdf))
                if rows:
                    all_rows.extend(rows)
                    self._q.put(("log",
                                 f"       ✓  {len(rows)} subject(s) extracted.", "ok"))
                else:
                    self._q.put(("log",
                                 "       ⚠  No subjects found — check scan quality.", "warn"))
            except Exception as exc:
                errors.append(pdf.name)
                self._q.put(("log", f"       ✗  Error: {exc}", "err"))

        self._q.put(("progress", 100))

        if all_rows:
            subjects = sorted({r.get("Subject Code", "") for r in all_rows if r.get("Subject Code", "")})
            self._q.put(("need_credits", {
                "rows": all_rows,
                "errors": errors,
                "total_files": n,
                "subjects": subjects,
            }))
        else:
            self._q.put(("log",      "\n✗  No data extracted from any PDF.", "err"))
            self._q.put(("done_err", None))

    def _export_worker(self, payload: dict, credits: dict[str, int]):
        try:
            out = _save_excel(payload["rows"], self._folder, credits)
            n = payload["total_files"]
            errors = payload["errors"]
            msg = (f"\nDone!  "
                   f"{len(payload['rows'])} row(s) from "
                   f"{n - len(errors)}/{n} file(s)\n"
                   f"   Saved: {out.name}")
            self._q.put(("log", msg, "done"))
            self._q.put(("done_ok", str(out)))
        except Exception as exc:
            self._q.put(("log", f"\n✗  Could not write Excel: {exc}", "err"))
            self._q.put(("done_err", None))

    def _prompt_subject_credits(self, subjects: list[str]) -> dict[str, int] | None:
        dlg = tk.Toplevel(self)
        dlg.title("Subject Credits")
        dlg.configure(bg=BG)
        dlg.transient(self)
        dlg.grab_set()
        dlg.resizable(False, False)

        frm = tk.Frame(dlg, bg=BG, padx=16, pady=12)
        frm.pack(fill="both", expand=True)

        tk.Label(
            frm,
            text="Enter credits for each subject (required before export):",
            font=F_HEAD,
            bg=BG,
            fg=T_DARK,
            anchor="w",
            justify="left",
        ).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 10))

        vars_by_code: dict[str, tk.StringVar] = {}
        for idx, code in enumerate(subjects, start=1):
            tk.Label(frm, text=code, font=F_BODY, bg=BG, fg=T_DARK).grid(
                row=idx, column=0, sticky="w", padx=(0, 12), pady=3
            )
            default = str(self._last_credits.get(code, ""))
            var = tk.StringVar(value=default)
            ent = tk.Entry(frm, textvariable=var, width=8, font=F_BODY)
            ent.grid(row=idx, column=1, sticky="w", pady=3)
            vars_by_code[code] = var
            if idx == 1:
                ent.focus_set()

        result: dict[str, int] | None = None

        def on_submit():
            nonlocal result
            credits: dict[str, int] = {}
            for code in subjects:
                raw = vars_by_code[code].get().strip()
                if raw == "":
                    messagebox.showerror("Missing Credit", f"Credit missing for {code}", parent=dlg)
                    return
                if not raw.isdigit():
                    messagebox.showerror("Invalid Credit", f"Credit must be a non-negative integer for {code}", parent=dlg)
                    return
                credits[code] = int(raw)
            result = credits
            dlg.destroy()

        def on_cancel():
            dlg.destroy()

        btns = tk.Frame(frm, bg=BG)
        btns.grid(row=len(subjects) + 1, column=0, columnspan=2, sticky="e", pady=(12, 0))
        tk.Button(btns, text="Cancel", font=F_BODY, command=on_cancel, width=10).pack(side="right")
        tk.Button(btns, text="Export", font=F_BODY, command=on_submit, width=10).pack(side="right", padx=(0, 8))

        dlg.protocol("WM_DELETE_WINDOW", on_cancel)
        self.wait_window(dlg)
        return result


    def _poll_queue(self):
        try:
            while True:
                item = self._q.get_nowait()
                kind = item[0]
                if kind == "log":
                    self._log_msg(item[1], item[2])
                elif kind == "progress":
                    self._progress_var.set(item[1])
                elif kind == "need_credits":
                    payload = item[1]
                    credits = self._prompt_subject_credits(payload["subjects"])
                    if not credits:
                        self._log_msg("\n✗  Export cancelled: credits were not provided.", "err")
                        self._busy = False
                        self._btn_parse.configure(state="normal",
                                                  text="  Parse & Export to Excel  ")
                        self._btn_browse.configure(state="normal")
                    else:
                        self._last_credits = credits
                        self._log_msg("Credits captured. Preparing workbook ...", "info")
                        threading.Thread(
                            target=self._export_worker,
                            args=(payload, credits),
                            daemon=True,
                        ).start()
                elif kind == "done_ok":
                    self._busy = False
                    self._btn_parse.configure(state="normal",
                                              text="  Parse & Export to Excel  ")
                    self._btn_browse.configure(state="normal")
                    messagebox.showinfo(
                        "Export Complete",
                        f"All results saved to:\n\n{item[1]}"
                    )
                elif kind == "done_err":
                    self._busy = False
                    self._btn_parse.configure(state="normal",
                                              text="  Parse & Export to Excel  ")
                    self._btn_browse.configure(state="normal")
        except queue.Empty:
            pass
        self.after(80, self._poll_queue)


    def _log_msg(self, text: str, tag: str = "info"):
        self._log.configure(state="normal")
        self._log.insert("end", text + "\n", tag)
        self._log.see("end")
        self._log.configure(state="disabled")


if __name__ == "__main__":
    app = VTUParserApp()
    app.mainloop()
